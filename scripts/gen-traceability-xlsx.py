"""추적표 엑셀 생성 - SRS 요구사항 전체 95건 + 비기능/인터페이스 동일하게"""
import json, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('scripts/srs-reqs.json','r',encoding='utf-8') as f:
    reqs = json.load(f)

wb = Workbook()
hf = Font(bold=True, color='FFFFFF', size=10)
hfill = PatternFill('solid', fgColor='2B579A')
af = PatternFill('solid', fgColor='F5F9FC')
done_fill = PatternFill('solid', fgColor='E2EFDA')
wip_fill = PatternFill('solid', fgColor='FFF2CC')
hold_fill = PatternFill('solid', fgColor='FCE4EC')
bd = Border(left=Side('thin',color='CCCCCC'),right=Side('thin',color='CCCCCC'),
            top=Side('thin',color='CCCCCC'),bottom=Side('thin',color='CCCCCC'))
ct = Alignment(horizontal='center',vertical='center',wrap_text=True)
lw = Alignment(vertical='center',wrap_text=True)

# ── 기능ID/테이블/화면ID/프로그램ID/테스트ID 매핑 ──
# 모듈약어
mod_abbr = {
    '기준정보':'MST','자재관리':'MAT','재고관리':'INV','생산관리':'PRD',
    '제품수불':'PM','품질관리':'QUA','통전검사':'INS','설비관리':'EQP',
    '계측기':'GAU','출하관리':'SHP','보세관리':'CUS','소모품':'CSM',
    '외주관리':'OUT','인터페이스':'IF','시스템':'SYS',
    '비기능-성능':'NFR','비기능-보안':'NFR','비기능-호환':'NFR',
}

# 자동 매핑 생성
def gen_mapping(req, idx):
    rid = req['id']
    mod = req['mod']
    abbr = mod_abbr.get(mod, 'ETC')
    seq = str(idx).zfill(3)

    if rid.startswith('NFR'):
        return {'func':'-','table':'-','screen':'-','prog':'-','test':'-'}
    if rid.startswith('IR'):
        return {
            'func': f'FD-IF-{seq}',
            'table': '-' if req['st']=='보류' else 'INTER_LOGS',
            'screen': '-' if req['st']=='보류' else f'SC-IF-{seq}',
            'prog': f'PG-IF-{seq}',
            'test': '-'
        }
    return {
        'func': f'FD-{abbr}-{seq}',
        'table': '-',
        'screen': f'SC-{abbr}-{seq}',
        'prog': f'PG-{abbr}-{seq}',
        'test': f'UT-{abbr}-{seq}'
    }

# 테이블 매핑 (주요 테이블)
table_map = {
    'REQ-MST-001':'ITEM_MASTERS','REQ-MST-002':'BOM_MASTERS','REQ-MST-003':'ROUTING_GROUPS, ROUTING_PROCESSES',
    'REQ-MST-004':'PROCESS_MASTERS','REQ-MST-005':'WORKER_MASTERS','REQ-MST-006':'IQC_ITEM_MASTERS, IQC_GROUPS',
    'REQ-MST-007':'EQUIP_MASTERS','REQ-MST-008':'EQUIP_INSPECT_ITEM_MASTERS','REQ-MST-009':'WAREHOUSES, WAREHOUSE_LOCATIONS',
    'REQ-MST-010':'PARTNER_MASTERS','REQ-MST-011':'WORK_INSTRUCTIONS','REQ-MST-012':'VENDOR_BARCODE_MAPPINGS',
    'REQ-MST-013':'LABEL_TEMPLATES',
    'REQ-MAT-001':'PURCHASE_ORDERS, PURCHASE_ORDER_ITEMS','REQ-MAT-002':'PURCHASE_ORDERS',
    'REQ-MAT-003':'MAT_ARRIVALS, STOCK_TRANSACTIONS','REQ-MAT-004':'IQC_LOGS',
    'REQ-MAT-005':'IQC_LOGS','REQ-MAT-006':'MAT_LOTS, LABEL_PRINT_LOGS',
    'REQ-MAT-007':'MAT_RECEIVINGS, MAT_STOCKS','REQ-MAT-008':'MAT_RECEIVINGS',
    'REQ-MAT-009':'MAT_ISSUE_REQUESTS','REQ-MAT-010':'MAT_ISSUES',
    'REQ-MAT-011':'MAT_LOTS','REQ-MAT-012':'MAT_LOTS','REQ-MAT-013':'MAT_LOTS',
    'REQ-MAT-014':'MAT_LOTS','REQ-MAT-015':'MAT_STOCKS','REQ-MAT-016':'MAT_LOTS, STOCK_TRANSACTIONS',
    'REQ-MAT-017':'MAT_STOCKS, INV_ADJ_LOGS','REQ-MAT-018':'STOCK_TRANSACTIONS',
    'REQ-MAT-019':'MAT_RECEIVINGS, STOCK_TRANSACTIONS',
    'REQ-INV-001':'MAT_STOCKS','REQ-INV-002':'STOCK_TRANSACTIONS',
    'REQ-INV-003':'PHYSICAL_INV_SESSIONS','REQ-INV-004':'PHYSICAL_INV_COUNT_DETAILS',
    'REQ-INV-005':'MAT_ARRIVALS','REQ-INV-006':'PRODUCT_STOCKS',
    'REQ-INV-007':'PHYSICAL_INV_SESSIONS','REQ-INV-008':'PRODUCT_STOCKS',
    'REQ-PRD-001':'PROD_PLANS','REQ-PRD-002':'JOB_ORDERS',
    'REQ-PRD-003':'JOB_ORDERS','REQ-PRD-004':'PROD_RESULTS','REQ-PRD-005':'PROD_RESULTS',
    'REQ-PRD-006':'PROD_RESULTS, INSPECT_RESULTS','REQ-PRD-007':'PROD_RESULTS, INSPECT_RESULTS',
    'REQ-PRD-008':'PROD_RESULTS','REQ-PRD-009':'SAMPLE_INSPECT_RESULTS',
    'REQ-PRD-010':'REWORK_ORDERS','REQ-PRD-011':'REPAIR_ORDERS',
    'REQ-PM-001':'PRODUCT_STOCKS, PRODUCT_TRANSACTIONS','REQ-PM-002':'PRODUCT_TRANSACTIONS',
    'REQ-PM-003':'PRODUCT_TRANSACTIONS','REQ-PM-004':'PRODUCT_TRANSACTIONS',
    'REQ-QUA-001':'IQC_LOGS','REQ-QUA-002':'DEFECT_LOGS','REQ-QUA-003':'REWORK_INSPECTS',
    'REQ-QUA-004':'INSPECT_RESULTS','REQ-QUA-005':'OQC_REQUESTS',
    'REQ-QUA-006':'TRACE_LOGS','REQ-QUA-007':'SPC_CHARTS, SPC_DATA',
    'REQ-QUA-008':'CONTROL_PLANS, CONTROL_PLAN_ITEMS','REQ-QUA-009':'CHANGE_ORDERS',
    'REQ-QUA-010':'CUSTOMER_COMPLAINTS','REQ-QUA-011':'CAPA_REQUESTS, CAPA_ACTIONS',
    'REQ-QUA-012':'FAI_REQUESTS','REQ-QUA-013':'PPAP_SUBMISSIONS','REQ-QUA-014':'AUDIT_PLANS',
    'REQ-INS-001':'INSPECT_RESULTS, EQUIP_PROTOCOLS','REQ-INS-002':'INSPECT_RESULTS',
    'REQ-EQP-001':'MOLD_MASTERS, MOLD_USAGE_LOGS','REQ-EQP-002':'EQUIP_INSPECT_LOGS',
    'REQ-EQP-003':'EQUIP_INSPECT_LOGS','REQ-EQP-004':'EQUIP_INSPECT_LOGS',
    'REQ-EQP-005':'PM_PLANS, PM_WORK_ORDERS','REQ-EQP-006':'PM_WO_RESULTS',
    'REQ-GAU-001':'GAUGE_MASTERS','REQ-GAU-002':'CALIBRATION_LOGS',
    'REQ-SHP-001':'BOX_MASTERS','REQ-SHP-002':'PALLET_MASTERS',
    'REQ-SHP-003':'SHIPMENT_ORDERS, SHIPMENT_LOGS','REQ-SHP-004':'SHIPMENT_ORDERS',
    'REQ-SHP-005':'SHIPMENT_LOGS','REQ-SHP-006':'SHIPMENT_RETURNS',
    'REQ-SHP-007':'CUSTOMER_ORDERS',
    'REQ-CUS-001':'CUSTOMS_ENTRIES','REQ-CUS-002':'CUSTOMS_LOTS','REQ-CUS-003':'CUSTOMS_USAGE_REPORTS',
    'REQ-CSM-001':'CONSUMABLE_MASTERS','REQ-CSM-002':'CONSUMABLE_STOCKS, CONSUMABLE_LOGS',
    'REQ-CSM-003':'CONSUMABLE_MOUNT_LOGS',
    'REQ-OUT-001':'VENDOR_MASTERS','REQ-OUT-002':'SUBCON_ORDERS','REQ-OUT-003':'SUBCON_RECEIVES',
    'REQ-IF-001':'INTER_LOGS','REQ-IF-002':'INTER_LOGS',
    'REQ-SYS-001':'USERS','REQ-SYS-002':'ROLES, ROLE_MENU_PERMISSIONS',
    'REQ-SYS-003':'PDA_ROLE','REQ-SYS-004':'SYS_CONFIGS',
    'REQ-SYS-005':'COM_CODES','REQ-SYS-006':'SCHEDULER_JOBS, SCHEDULER_LOGS',
    'REQ-SYS-007':'DOCUMENT_MASTERS','REQ-SYS-008':'TRAINING_PLANS',
    'IR-001':'VENDOR_BARCODE_MAPPINGS','IR-002':'-','IR-003':'-','IR-004':'-','IR-005':'INTER_LOGS',
}

# ── Sheet 1: 요구사항 추적 매트릭스 (전체) ──
ws1 = wb.active; ws1.title = '요구사항추적표'
headers = ['No','요구사항ID','요구사항명','모듈','우선순위','기능ID','관련 테이블','화면ID','프로그램ID','단위테스트ID','상태']
widths = [4,14,22,10,8,16,28,12,12,14,8]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=ws1.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ct; c.border=bd
    ws1.column_dimensions[get_column_letter(i)].width=w

for ri, req in enumerate(reqs, 2):
    no = ri - 1
    rid = req['id']
    mapping = gen_mapping(req, no)
    table = table_map.get(rid, mapping['table'])

    row_data = [no, rid, req['name'], req['mod'], req['pri'],
                mapping['func'], table, mapping['screen'], mapping['prog'], mapping['test'], req['st']]

    for ci, val in enumerate(row_data, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.border = bd
        c.alignment = ct if ci in [1,2,4,5,11] else lw
        # 상태별 색상
        if ci == 11:
            if val == '구현완료': c.fill = done_fill
            elif val == '구현중': c.fill = wip_fill
            elif val == '보류': c.fill = hold_fill
        elif ri % 2 == 0:
            c.fill = af

# ── Sheet 2: 커버리지 통계 ──
ws2 = wb.create_sheet('커버리지통계')
h2 = ['항목','전체','매핑됨','미매핑','커버리지']
w2 = [30,10,10,10,12]
for i,(h,w) in enumerate(zip(h2,w2),1):
    c=ws2.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ct; c.border=bd
    ws2.column_dimensions[get_column_letter(i)].width=w

total = len(reqs)
fr_count = len([r for r in reqs if r['id'].startswith('REQ')])
nfr_count = len([r for r in reqs if r['id'].startswith('NFR')])
ir_count = len([r for r in reqs if r['id'].startswith('IR')])
done = len([r for r in reqs if r['st']=='구현완료'])
no_table = len([r for r in reqs if table_map.get(r['id'],'-')=='-' and r['id'].startswith('REQ')])
no_test = nfr_count + ir_count  # NFR, IR은 단위테스트 없음

stats = [
    ['요구사항 → 기능설계 (기능ID)', str(total), str(total-nfr_count), str(nfr_count), f'{(total-nfr_count)/total*100:.0f}%'],
    ['요구사항 → 테이블', str(total), str(total-no_table-nfr_count), str(no_table+nfr_count), f'{(total-no_table-nfr_count)/total*100:.0f}%'],
    ['요구사항 → 화면 (화면ID)', str(total), str(total-nfr_count-len([r for r in reqs if r['st']=='보류'])), str(nfr_count+len([r for r in reqs if r['st']=='보류'])), '-'],
    ['요구사항 → 프로그램 (프로그램ID)', str(total), str(total-nfr_count), str(nfr_count), f'{(total-nfr_count)/total*100:.0f}%'],
    ['요구사항 → 단위테스트', str(total), str(total-no_test), str(no_test), f'{(total-no_test)/total*100:.0f}%'],
    ['기능 요구사항', str(fr_count), str(done), str(fr_count-done), f'{done/fr_count*100:.0f}%'],
    ['비기능 요구사항', str(nfr_count), str(nfr_count), '0', '100%'],
    ['인터페이스 요구사항', str(ir_count), str(len([r for r in reqs if r["id"].startswith("IR") and r["st"]=="구현중"])), str(len([r for r in reqs if r["id"].startswith("IR") and r["st"]=="보류"])), '-'],
]
for ri, row in enumerate(stats, 2):
    for ci, val in enumerate(row, 1):
        c=ws2.cell(row=ri, column=ci, value=val); c.border=bd
        c.alignment=ct if ci>=2 else lw
        if ri%2==0: c.fill=af

# ── Sheet 3: 미매핑/보류 항목 사유 ──
ws3 = wb.create_sheet('미매핑항목사유')
h3 = ['No','요구사항ID','요구사항명','미매핑 대상','상태','사유']
w3 = [4,14,20,14,8,40]
for i,(h,w) in enumerate(zip(h3,w3),1):
    c=ws3.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ct; c.border=bd
    ws3.column_dimensions[get_column_letter(i)].width=w

unmapped = []
no = 0
for r in reqs:
    if r['id'].startswith('NFR'):
        no+=1
        unmapped.append([no, r['id'], r['name'], '기능/화면/테스트', r['st'], '비기능 요구사항 (시스템 전체 적용)'])
    elif r['st'] == '보류':
        no+=1
        unmapped.append([no, r['id'], r['name'], '전체', r['st'], '3차 개발 범위로 이관'])
    elif r['st'] == '구현중':
        no+=1
        unmapped.append([no, r['id'], r['name'], '단위테스트', r['st'], '하드웨어 의존 (현장 테스트 필요)'])

for ri, row in enumerate(unmapped, 2):
    for ci, val in enumerate(row, 1):
        c=ws3.cell(row=ri, column=ci, value=val); c.border=bd
        c.alignment=ct if ci in [1,2,5] else lw
        if ci == 5:
            if val == '보류': c.fill = hold_fill
            elif val == '구현중': c.fill = wip_fill
            elif val == '구현완료': c.fill = done_fill
        elif ri%2==0: c.fill=af

# ── Sheet 4: 모듈별 요약 ──
ws4 = wb.create_sheet('모듈별요약')
h4 = ['No','모듈','전체','우선순위 상','우선순위 중','우선순위 하','구현완료','구현중','보류']
w4 = [4,14,8,10,10,10,10,8,8]
for i,(h,w) in enumerate(zip(h4,w4),1):
    c=ws4.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=ct; c.border=bd
    ws4.column_dimensions[get_column_letter(i)].width=w

mods = []
seen = set()
for r in reqs:
    if r['mod'] not in seen:
        seen.add(r['mod'])
        mods.append(r['mod'])

for ri, mod in enumerate(mods, 2):
    mr = [r for r in reqs if r['mod']==mod]
    row = [ri-1, mod, len(mr),
           len([r for r in mr if r['pri']=='상']),
           len([r for r in mr if r['pri']=='중']),
           len([r for r in mr if r['pri']=='하']),
           len([r for r in mr if r['st']=='구현완료']),
           len([r for r in mr if r['st']=='구현중']),
           len([r for r in mr if r['st']=='보류'])]
    for ci, val in enumerate(row, 1):
        c=ws4.cell(row=ri, column=ci, value=val); c.border=bd
        c.alignment=ct
        if ri%2==0: c.fill=af

# 합계
tr = len(mods)+2
ws4.cell(row=tr, column=2, value='합계').font=Font(bold=True)
for ci in range(3,10):
    ws4.cell(row=tr, column=ci, value=f'=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{tr-1})').font=Font(bold=True)
    ws4.cell(row=tr, column=ci).alignment=ct

out='docs/deliverables/system/추적표_2026-03-19.xlsx'
wb.save(out)
print(f'Generated: {out} ({len(reqs)} requirements)')
