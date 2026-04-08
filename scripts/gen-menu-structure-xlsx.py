"""
HANES MES 메뉴구성도 Excel 생성
실행: python scripts/gen-menu-structure-xlsx.py
출력: exports/all/메뉴구성도_YYYY-MM-DD.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

today = date.today().isoformat()
wb = Workbook()

# ── 색상/스타일 ──
hdr_fill = PatternFill("solid", fgColor="1F4E79")
hdr_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
sub_fill = PatternFill("solid", fgColor="D6E4F0")
sub_font = Font(name="맑은 고딕", bold=True, size=11)
data_font = Font(name="맑은 고딕", size=10)
bold_font = Font(name="맑은 고딕", bold=True, size=10)
title_font = Font(name="맑은 고딕", bold=True, size=16, color="1F4E79")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 메뉴 데이터 (menuConfig + i18n ko.json) ──
menus = [
    ("DASHBOARD", "대시보드", None, []),
    ("MONITORING", "모니터링", None, [
        ("MON_EQUIP_STATUS", "설비 가동현황", "/equipment/status"),
    ]),
    ("MASTER", "기준정보", None, [
        ("MST_PART", "품목관리", "/master/part"),
        ("MST_BOM", "BOM관리", "/master/bom"),
        ("MST_PARTNER", "거래처관리", "/master/partner"),
        ("EQUIP_MASTER", "설비마스터", "/master/equip"),
        ("MST_PROCESS", "공정관리", "/master/process"),
        ("MST_PROD_LINE", "생산라인관리", "/master/prod-line"),
        ("MST_ROUTING", "라우팅관리", "/master/routing"),
        ("MST_WORKER", "작업자관리", "/master/worker"),
        ("MST_WORK_INST", "작업지도서관리", "/master/work-instruction"),
        ("MST_WAREHOUSE", "창고관리", "/master/warehouse"),
        ("MST_LABEL", "라벨관리", "/master/label"),
        ("MST_VENDOR_BARCODE", "제조사 바코드 매핑", "/master/vendor-barcode"),
        ("SYS_DOCUMENT", "문서관리", "/system/document"),
    ]),
    ("INVENTORY", "자재재고관리", None, [
        ("INV_MAT_STOCK", "자재재고현황조회", "/inventory/material-stock"),
        ("INV_TRANSACTION", "자재수불이력조회", "/inventory/transaction"),
        ("INV_MAT_PHYSICAL_INV", "자재재고실사관리", "/inventory/material-physical-inv"),
        ("INV_MAT_PHYSICAL_INV_HISTORY", "자재재고실사조회", "/inventory/material-physical-inv-history"),
        ("INV_ARRIVAL_STOCK", "입하재고현황조회", "/material/arrival-stock"),
        ("MAT_HOLD", "자재재고홀드관리", "/material/hold"),
    ]),
    ("PRODUCT_INVENTORY", "제품재고관리", None, [
        ("INV_PRODUCT_STOCK", "제품재고현황조회", "/inventory/stock"),
        ("INV_PRODUCT_PHYSICAL_INV", "제품재고실사관리", "/inventory/product-physical-inv"),
        ("INV_PRODUCT_PHYSICAL_INV_HISTORY", "제품재고실사조회", "/inventory/product-physical-inv-history"),
        ("PROD_HOLD", "제품재고홀드관리", "/inventory/product-hold"),
    ]),
    ("PRODUCT_MGMT", "제품수불관리", None, [
        ("PROD_RECEIVE", "제품입고관리", "/product/receive"),
        ("PROD_RECEIPT_CANCEL", "제품입고취소", "/product/receipt-cancel"),
        ("PROD_ISSUE", "제품출고관리", "/product/issue"),
        ("PROD_ISSUE_CANCEL", "제품출고취소", "/product/issue-cancel"),
    ]),
    ("MATERIAL", "자재수불관리", None, [
        ("MAT_ARRIVAL", "입하관리", "/material/arrival"),
        ("MAT_RECEIVE_LABEL", "입고라벨발행", "/material/receive-label"),
        ("MAT_RECEIVE", "자재입고관리", "/material/receive"),
        ("MAT_RECEIVE_HISTORY", "입고이력조회", "/material/receive-history"),
        ("MAT_REQUEST", "출고요청관리", "/material/request"),
        ("MAT_ISSUE", "자재출고관리", "/material/issue"),
        ("MAT_LOT", "LOT관리", "/material/lot"),
        ("MAT_LOT_SPLIT", "자재분할관리", "/material/lot-split"),
        ("MAT_LOT_MERGE", "자재병합관리", "/material/lot-merge"),
        ("MAT_SHELF_LIFE", "자재유수명관리", "/material/shelf-life"),
        ("MAT_SCRAP", "자재폐기처리", "/material/scrap"),
        ("MAT_ADJUSTMENT", "재고보정처리", "/material/adjustment"),
        ("MAT_MISC_RECEIPT", "기타입고관리", "/material/misc-receipt"),
        ("MAT_RECEIPT_CANCEL", "자재입고취소", "/material/receipt-cancel"),
    ]),
    ("PURCHASING", "자재주문관리", None, [
        ("PUR_PO", "PO관리", "/material/po"),
        ("PUR_PO_STATUS", "PO현황조회", "/material/po-status"),
    ]),
    ("PRODUCTION", "생산관리", None, [
        ("PROD_MONTHLY_PLAN", "월간생산계획", "/production/monthly-plan"),
        ("PROD_ORDER", "작업지시관리", "/production/order"),
        ("PROD_RESULT", "생산실적조회", "/production/result"),
        ("PROD_PROGRESS", "작업지시현황조회", "/production/progress"),
        ("PROD_INPUT_MANUAL", "실적입력(수작업)", "/production/input-manual"),
        ("PROD_INPUT_MACHINE", "실적입력(가공)", "/production/input-machine"),
        ("PROD_INPUT_INSPECT", "실적입력(단순검사)", "/production/input-inspect"),
        ("PROD_INPUT_EQUIP", "실적입력(검사장비)", "/production/input-equip"),
        ("PROD_RESULT_SUMMARY", "작업실적통합조회", "/production/result-summary"),
        ("PROD_WIP_STOCK", "반제품/제품재고조회", "/production/wip-stock"),
        ("QC_REWORK", "재작업 지시", "/quality/rework"),
        ("QC_REWORK_HISTORY", "재작업 현황", "/quality/rework-history"),
        ("PROD_REPAIR", "수리관리", "/production/repair"),
    ]),
    ("INSPECTION", "통전검사", None, [
        ("INSP_RESULT", "통전검사관리", "/inspection/result"),
        ("INSP_HISTORY", "통전검사이력", "/inspection/history"),
        ("INSP_PROTOCOL", "검사기프로토콜", "/inspection/protocol"),
    ]),
    ("QUALITY", "품질관리", None, [
        ("QC_IQC_ITEM", "IQC검사항목/그룹", "/master/iqc-item"),
        ("QC_IQC", "수입검사(IQC)", "/material/iqc"),
        ("QC_IQC_HISTORY", "수입검사이력조회", "/material/iqc-history"),
        ("QC_DEFECT", "불량등록관리", "/quality/defect"),
        ("QC_REWORK_INSPECT", "재작업 후 검사", "/quality/rework-inspect"),
        ("QC_INSPECT", "외관검사", "/quality/inspect"),
        ("QC_SAMPLE_INSPECT", "샘플검사이력조회", "/production/sample-inspect"),
        ("QC_OQC", "출하검사(OQC)", "/quality/oqc"),
        ("QC_OQC_HISTORY", "출하검사이력조회", "/quality/oqc-history"),
        ("QC_TRACE", "추적성조회", "/quality/trace"),
        ("QC_CHANGE", "변경점관리", "/quality/change-control"),
        ("QC_COMPLAINT", "고객클레임", "/quality/complaint"),
        ("QC_CAPA", "CAPA관리", "/quality/capa"),
        ("QC_FAI", "초물검사", "/quality/fai"),
        ("QC_PPAP", "PPAP관리", "/quality/ppap"),
        ("QC_SPC", "SPC관리", "/quality/spc"),
        ("QC_CONTROL_PLAN", "관리계획서", "/quality/control-plan"),
        ("QC_AUDIT", "내부심사", "/quality/audit"),
        ("SYS_TRAINING", "교육훈련", "/system/training"),
    ]),
    ("EQUIPMENT", "설비관리", None, [
        ("EQ_MOLD_MGMT", "금형관리", "/equipment/mold-mgmt"),
        ("EQUIP_INSPECT_ITEM_MASTER", "점검항목마스터", "/master/equip-inspect-item"),
        ("EQUIP_INSPECT_ITEM", "설비별 점검항목", "/master/equip-inspect"),
        ("EQUIP_INSPECT_CALENDAR", "일상점검 캘린더", "/equipment/inspect-calendar"),
        ("EQUIP_DAILY", "일상점검 결과", "/equipment/daily-inspect"),
        ("EQUIP_PERIODIC_CALENDAR", "정기점검 캘린더", "/equipment/periodic-inspect-calendar"),
        ("EQUIP_PERIODIC", "정기점검 결과", "/equipment/periodic-inspect"),
        ("EQUIP_HISTORY", "점검이력 조회", "/equipment/inspect-history"),
        ("EQUIP_PM_PLAN", "PM 계획", "/equipment/pm-plan"),
        ("EQUIP_PM_CALENDAR", "PM 캘린더", "/equipment/pm-calendar"),
        ("EQUIP_PM_RESULT", "PM 보전결과", "/equipment/pm-result"),
    ]),
    ("GAUGE_MGMT", "계측기관리", None, [
        ("GAUGE_MASTER", "계측기마스터", "/master/gauge"),
        ("GAUGE_CALIBRATION", "계측기 교정관리", "/quality/msa"),
        ("GAUGE_CALIBRATION_HISTORY", "교정이력 조회", "/equipment/calibration-history"),
    ]),
    ("SHIPPING", "출하관리", None, [
        ("SHIP_PACK_RESULT", "포장실적조회", "/production/pack-result"),
        ("SHIP_PACK", "제품포장관리", "/shipping/pack"),
        ("SHIP_PALLET", "팔레트적재관리", "/shipping/pallet"),
        ("SHIP_CONFIRM", "출하확정처리", "/shipping/confirm"),
        ("SHIP_ORDER", "출하지시등록", "/shipping/order"),
        ("SHIP_HISTORY", "출하이력조회", "/shipping/history"),
        ("SHIP_RETURN", "출하반품등록", "/shipping/return"),
        ("SHIP_CUST_PO", "고객발주관리", "/shipping/customer-po"),
        ("SHIP_CUST_PO_STATUS", "고객발주현황조회", "/shipping/customer-po-status"),
    ]),
    ("CUSTOMS", "보세관리", None, [
        ("CUST_ENTRY", "수입신고관리", "/customs/entry"),
        ("CUST_STOCK", "보세재고조회", "/customs/stock"),
        ("CUST_USAGE", "사용신고관리", "/customs/usage"),
    ]),
    ("CONSUMABLES", "소모품관리", None, [
        ("CONS_MASTER", "소모품관리", "/consumables/master"),
        ("CONS_LABEL", "라벨발행", "/consumables/label"),
        ("CONS_RECEIVING", "입고관리", "/consumables/receiving"),
        ("CONS_ISSUING", "출고관리", "/consumables/issuing"),
        ("CONS_STOCK", "재고현황조회", "/consumables/stock"),
        ("CONS_LIFE", "수명현황조회", "/consumables/life"),
    ]),
    ("OUTSOURCING", "외주관리", None, [
        ("OUT_VENDOR", "외주처관리", "/outsourcing/vendor"),
        ("OUT_ORDER", "외주발주관리", "/outsourcing/order"),
        ("OUT_RECEIVE", "외주입고관리", "/outsourcing/receive"),
    ]),
    ("INTERFACE", "인터페이스", None, [
        ("IF_DASHBOARD", "ERP연동현황조회", "/interface/dashboard"),
        ("IF_LOG", "전송이력조회", "/interface/log"),
        ("IF_MANUAL", "수동전송처리", "/interface/manual"),
    ]),
    ("SYSTEM", "시스템관리", None, [
        ("SYS_COMPANY", "회사관리", "/master/company"),
        ("SYS_DEPT", "부서관리", "/system/department"),
        ("SYS_USER", "사용자관리", "/system/users"),
        ("SYS_ROLE", "역할관리", "/system/roles"),
        ("SYS_PDA_ROLE", "PDA 역할관리", "/system/pda-roles"),
        ("SYS_COMM", "통신설정관리", "/system/comm-config"),
        ("SYS_CONFIG", "환경설정", "/system/config"),
        ("SYS_CODE", "코드관리", "/master/code"),
        ("SYS_SCHEDULER", "스케줄러관리", "/system/scheduler"),
    ]),
]

# ═══════════════════════════════════════════════════════════════
# Sheet 1: 메뉴구성도 (트리형)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "메뉴구성도"
ws1.sheet_properties.tabColor = "1F4E79"

# 타이틀
ws1.merge_cells("A1:F1")
ws1["A1"] = "HANES MES 메뉴구성도"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:F2")
ws1["A2"] = f"작성일: {today}  |  총 메뉴: {sum(1 + len(m[3]) for m in menus)}개 (대메뉴 {len(menus)}개)"
ws1["A2"].font = Font(name="맑은 고딕", size=10, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[2].height = 22

# 헤더
headers = ["No", "대메뉴", "소메뉴", "메뉴코드(RBAC)", "라우트 경로", "비고"]
col_widths = [6, 18, 24, 28, 32, 15]
for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(c)].width = w

# 데이터
row = 5
no = 0
total_sub = 0
for parent_code, parent_name, parent_path, children in menus:
    no += 1
    # 대메뉴 행
    cell_no = ws1.cell(row=row, column=1, value=no)
    cell_parent = ws1.cell(row=row, column=2, value=parent_name)
    cell_sub = ws1.cell(row=row, column=3, value="-" if children else parent_name)
    cell_code = ws1.cell(row=row, column=4, value=parent_code)
    cell_path = ws1.cell(row=row, column=5, value=parent_path or "(하위 메뉴 포함)")
    cell_note = ws1.cell(row=row, column=6, value=f"하위 {len(children)}개" if children else "단독 메뉴")

    for c in range(1, 7):
        cell = ws1.cell(row=row, column=c)
        cell.font = sub_font
        cell.fill = sub_fill
        cell.border = thin_border
        cell.alignment = center if c in (1, 4, 6) else left_wrap
    row += 1

    # 소메뉴 행
    for ci, (child_code, child_name, child_path) in enumerate(children):
        total_sub += 1
        ws1.cell(row=row, column=1, value="").border = thin_border
        ws1.cell(row=row, column=2, value="").border = thin_border

        cell_sub = ws1.cell(row=row, column=3, value=f"  {child_name}")
        cell_sub.font = data_font
        cell_sub.border = thin_border
        cell_sub.alignment = left_wrap

        cell_code = ws1.cell(row=row, column=4, value=child_code)
        cell_code.font = data_font
        cell_code.border = thin_border
        cell_code.alignment = center

        cell_path = ws1.cell(row=row, column=5, value=child_path)
        cell_path.font = data_font
        cell_path.border = thin_border
        cell_path.alignment = left_wrap

        ws1.cell(row=row, column=6, value="").border = thin_border
        row += 1

# 합계 행
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws1.cell(row=row, column=1, value="합계").font = Font(name="맑은 고딕", bold=True, size=11)
ws1.cell(row=row, column=1).alignment = center
ws1.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
ws1.cell(row=row, column=1).border = thin_border
ws1.cell(row=row, column=2).border = thin_border
ws1.cell(row=row, column=3, value=f"대메뉴 {len(menus)}개 + 소메뉴 {total_sub}개 = 총 {len(menus) + total_sub}개")
ws1.cell(row=row, column=3).font = bold_font
ws1.cell(row=row, column=3).border = thin_border
for c in range(4, 7):
    ws1.cell(row=row, column=c).border = thin_border
    ws1.cell(row=row, column=c).fill = PatternFill("solid", fgColor="F2F2F2")

# ═══════════════════════════════════════════════════════════════
# Sheet 2: 통계 요약
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("통계요약")
ws2.sheet_properties.tabColor = "2E75B6"

ws2.merge_cells("A1:C1")
ws2["A1"] = "메뉴 통계 요약"
ws2["A1"].font = title_font
ws2["A1"].alignment = center
ws2.row_dimensions[1].height = 30

stat_headers = ["대메뉴", "소메뉴 수", "비율(%)"]
for c, h in enumerate(stat_headers, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = center
    cell.border = thin_border

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12

for i, (_, name, _, children) in enumerate(menus):
    r = 4 + i
    ws2.cell(row=r, column=1, value=name).font = data_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value=max(len(children), 1)).font = data_font
    ws2.cell(row=r, column=2).alignment = center
    ws2.cell(row=r, column=2).border = thin_border
    formula_pct = f"=B{r}/B{4+len(menus)}*100"
    ws2.cell(row=r, column=3, value=formula_pct).font = data_font
    ws2.cell(row=r, column=3).alignment = center
    ws2.cell(row=r, column=3).number_format = "0.0"
    ws2.cell(row=r, column=3).border = thin_border

total_row = 4 + len(menus)
ws2.cell(row=total_row, column=1, value="합계").font = bold_font
ws2.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
ws2.cell(row=total_row, column=1).border = thin_border
ws2.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})").font = bold_font
ws2.cell(row=total_row, column=2).alignment = center
ws2.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="F2F2F2")
ws2.cell(row=total_row, column=2).border = thin_border
ws2.cell(row=total_row, column=3, value="100.0").font = bold_font
ws2.cell(row=total_row, column=3).alignment = center
ws2.cell(row=total_row, column=3).fill = PatternFill("solid", fgColor="F2F2F2")
ws2.cell(row=total_row, column=3).border = thin_border

# ── 저장 ──
import os
out_dir = os.path.join(os.path.dirname(__file__), "..", "docs", "deliverables", "all")
os.makedirs(out_dir, exist_ok=True)
out_file = os.path.join(out_dir, f"메뉴구성도_{today}.xlsx")
wb.save(out_file)
print(f"생성 완료: {out_file}")
print(f"  대메뉴: {len(menus)}개, 소메뉴: {total_sub}개, 총: {len(menus) + total_sub}개")
